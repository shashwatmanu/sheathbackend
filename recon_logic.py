import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
import os

# === Validation helpers (approved) ===
def _require_exists(path: str, what: str):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"{what} not found: {p}")
    if not p.is_file():
        raise ValueError(f"{what} is not a file: {p}")
    return p

def _require_suffix(p: Path, allowed_suffixes, what: str):
    if p.suffix.lower() not in [s.lower() for s in allowed_suffixes]:
        raise ValueError(f"{what} must be one of {allowed_suffixes}, got: {p.suffix}")
    return p

def _require_parent_writable(path: str, what: str):
    parent = Path(path).parent
    if not parent.exists():
        raise FileNotFoundError(f"Parent directory for {what} does not exist: {parent}")
    if not os.access(parent, os.W_OK):
        raise PermissionError(f"Parent directory for {what} is not writable: {parent}")

# === Safe parsers (as in your script) ===
def safe_upper_strip(x):
    return str(x).strip().upper() if pd.notna(x) else ""

def safe_float_val(x):
    try:
        if pd.isna(x) or str(x).strip() == '':
            return None
        return float(str(x).replace(',', '').strip())
    except:
        return None

# === MAIN callable: exactly your flow, parameterized ===
def run_recon_pipeline(
    pdf_path: str,
    bank_file_paths: list[str],
    mis_file_path: str,
    outstanding_file_path: str,
    consolidated_output_path: str,
    updated_outstanding_path: str,
):
    # ------------ Columns (same as original) ------------
    bank_cols = ["Transaction ID", "Description", "Transaction Amount(INR)"]
    mis_cols = [
        "IHX Ref Id", "Hospital Name", "RohiniId", "Patient Name", "In Patient Number",
        "Claim Number", "Initial Claim Number", "Settled Amount", "TDS Amount",
        "Cheque/ NEFT/ UTR No.", "Cheque/ NEFT/ UTR Date", "Claim Status", "TPA Name"
    ]

    # ------------ Validations (approved) ------------
    _pdf_p = _require_exists(pdf_path, "PDF path")
    _require_suffix(_pdf_p, [".pdf"], "PDF path")

    if not isinstance(bank_file_paths, (list, tuple)) or len(bank_file_paths) == 0:
        raise ValueError("bank_file_paths must be a non-empty list/tuple of file paths.")
    for _bf in bank_file_paths:
        _bf_p = _require_exists(_bf, "Bank statement")
        _require_suffix(_bf_p, [".xlsx", ".xls", ".xlsm"], "Bank statement")

    _mis_p = _require_exists(mis_file_path, "MIS file")
    _require_suffix(_mis_p, [".xlsx", ".xls", ".xlsm"], "MIS file")

    _outstanding_p = _require_exists(outstanding_file_path, "Outstanding report file")
    _require_suffix(_outstanding_p, [".xlsx", ".xls", ".xlsm"], "Outstanding report file")

    _require_parent_writable(consolidated_output_path, "Consolidated output")
    _require_parent_writable(updated_outstanding_path, "Updated outstanding output")

    # ------------ 1) PDF -> DataFrame (same logic) ------------
    all_rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                all_rows.extend(table)

    if not all_rows:
        raise ValueError("No tables found in PDF.")

    columns = all_rows[0]
    rows = all_rows[1:]
    pdf_df = pd.DataFrame(rows, columns=columns)

    # normalize with underscores (matches original expectations)
    pdf_df.columns = [str(c).strip().replace('.', '_').replace(' ', '_') for c in pdf_df.columns]
    required_cols = ['Msg_Refer_No', 'Refer_No']
    if not all(col in pdf_df.columns for col in required_cols):
        raise ValueError(f"Missing one or more required columns: {required_cols}. Got: {list(pdf_df.columns)}")

    # Filter /XUTR/ & strip prefix
    pdf_df = pdf_df[pdf_df['Refer_No'].notna() & pdf_df['Refer_No'].str.strip().ne("")].copy()
    pdf_df = pdf_df[pdf_df['Refer_No'].str.contains("/XUTR/", na=False)]
    pdf_df['Refer_No'] = pdf_df['Refer_No'].str.replace('^/XUTR/', '', regex=True)

    # ------------ 2) Bank combine (same logic) ------------
    bank_dfs = [pd.read_excel(f, dtype=str)[bank_cols] for f in bank_file_paths]
    bank_df = pd.concat(bank_dfs, ignore_index=True)
    bank_df["Description"] = bank_df["Description"].astype(str)

    # ------------ 3) Match Msg_Refer_No in Bank Description ------------
    unique_pdf_msg_refs = pdf_df['Msg_Refer_No'].dropna().unique().tolist()
    matched_pdf_chunks = []
    for msg_ref in unique_pdf_msg_refs:
        if pd.isna(msg_ref) or str(msg_ref).strip() == '':
            continue
        mask = bank_df["Description"].str.contains(str(msg_ref), regex=False, na=False)
        if mask.any():
            tmp = bank_df.loc[mask, bank_cols].copy()
            tmp["__Msg_Ref__"] = msg_ref
            matched_pdf_chunks.append(tmp)

    if not matched_pdf_chunks:
        consolidated_result = pd.DataFrame(columns=bank_cols + mis_cols)
    else:
        pdf_hits_df = pd.concat(matched_pdf_chunks, ignore_index=True)
        merged_pdf = pdf_hits_df.merge(pdf_df, left_on="__Msg_Ref__", right_on="Msg_Refer_No", how="inner")

        # ------------ 4) Map Refer_No to MIS Cheque/ NEFT/ UTR No. ------------
        mis_df = pd.read_excel(mis_file_path, usecols=mis_cols, dtype=str)
        mis_clean = mis_df[
            mis_df["Cheque/ NEFT/ UTR No."].notna() & mis_df["Cheque/ NEFT/ UTR No."].str.strip().ne("")
        ].copy()

        consolidated_result = merged_pdf.merge(
            mis_clean,
            left_on="Refer_No",
            right_on="Cheque/ NEFT/ UTR No.",
            how="inner"
        )

        # Write consolidated excel
        pd.DataFrame(consolidated_result).to_excel(consolidated_output_path, index=False)

    # ------------ 5) Outstanding update (AND logic, same) ------------
    inpat_to_claim = {
        safe_upper_strip(row["In Patient Number"]): row["Claim Number"]
        for _, row in consolidated_result.iterrows()
    }
    patnm_to_claim = {
        safe_upper_strip(row["Patient Name"]): row["Claim Number"]
        for _, row in consolidated_result.iterrows()
    }
    claim_to_settled_amount = {
        safe_upper_strip(row["Claim Number"]): safe_float_val(row["Settled Amount"])
        for _, row in consolidated_result.iterrows()
    }

    cyan_fill = PatternFill(fill_type="solid", fgColor="00FFFF")
    update_found = False
    wb = openpyxl.load_workbook(outstanding_file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df = pd.read_excel(outstanding_file_path, sheet_name=sheet_name, dtype=str)
        if "Claim No" not in df.columns or "CR No" not in df.columns or "Patient Name" not in df.columns or "Balance" not in df.columns:
            continue

        for idx, row in df[df["Claim No"].isna() | (df["Claim No"].str.strip() == "")].iterrows():
            crno = safe_upper_strip(row["CR No"])
            patnm = safe_upper_strip(row["Patient Name"])
            balance = safe_float_val(row["Balance"])

            match_claim = None
            if (crno in inpat_to_claim) and (patnm in patnm_to_claim):
                claim_cr = inpat_to_claim[crno]
                claim_pat = patnm_to_claim[patnm]
                if claim_cr == claim_pat:
                    settled_amt = claim_to_settled_amount.get(claim_cr)
                    if settled_amt is not None and balance is not None and abs(settled_amt - balance) < 0.01:
                        match_claim = claim_cr

            if match_claim:
                update_found = True
                excel_row = idx + 2  # header is row 1
                header = [cell.value for cell in ws[1]]
                claim_col_idx = header.index("Claim No") + 1
                ws.cell(row=excel_row, column=claim_col_idx, value=match_claim)
                ws.cell(row=excel_row, column=claim_col_idx).fill = cyan_fill

    if update_found:
        wb.save(updated_outstanding_path)

    return {
        "matches_found": int(len(consolidated_result)) if not consolidated_result.empty else 0,
        "consolidated_written": consolidated_output_path if os.path.exists(consolidated_output_path) else None,
        "updated_outstanding_written": updated_outstanding_path if (update_found and os.path.exists(updated_outstanding_output := updated_outstanding_path) and os.path.exists(updated_outstanding_output)) else None,
    }
